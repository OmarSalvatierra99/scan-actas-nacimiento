"""
Servicio de generación de archivos Excel con datos de actas.
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import Config
from utils.logger import get_app_logger

logger = get_app_logger()


class ExcelGenerator:
    """Genera archivos Excel con los registros de actas escaneadas."""

    @staticmethod
    def generar_excel(registros: List[Dict[str, str]]) -> BytesIO:
        """
        Genera un archivo Excel en memoria con los registros.

        Args:
            registros: Lista de diccionarios con los datos de las actas

        Returns:
            BytesIO con el archivo Excel generado

        Raises:
            Exception: Si hay error al generar el Excel
        """
        try:
            logger.info(f"Generando Excel con {len(registros)} registro(s)")

            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Actas Escaneadas"

            # Estilo para encabezados
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Escribir encabezados
            for col, encabezado in enumerate(Config.ENCABEZADOS, 1):
                cell = ws.cell(row=1, column=col, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            logger.debug(f"Encabezados escritos: {len(Config.ENCABEZADOS)} columnas")

            # Escribir datos
            for row_idx, registro in enumerate(registros, 2):
                for col_idx, encabezado in enumerate(Config.ENCABEZADOS, 1):
                    valor = registro.get(encabezado, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    cell.alignment = Alignment(vertical="center")

            # Ajustar ancho de columnas
            ExcelGenerator._ajustar_anchos_columnas(ws)

            # Guardar en memoria
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            logger.info(f"Excel generado exitosamente ({bio.getbuffer().nbytes} bytes)")

            return bio

        except Exception as e:
            logger.error(f"Error al generar Excel: {e}", exc_info=True)
            raise

    @staticmethod
    def _ajustar_anchos_columnas(ws):
        """
        Ajusta automáticamente el ancho de las columnas.

        Args:
            ws: Worksheet de openpyxl
        """
        try:
            # Anchos predefinidos por columna
            anchos = {
                'Tomo': 8,
                'Libro': 8,
                'Foja': 8,
                'Acta': 10,
                'Entidad': 15,
                'Municipio': 20,
                'CURP': 20,
                'Registrado': 35,
                'Padre': 30,
                'Madre': 30,
                'FechaNacimiento': 15,
                'Sexo': 6,
                'FechaRegistro': 15,
                'Oficial': 10,
                'Folio': 15,
                'FechaEscaneo': 18
            }

            for col_idx, encabezado in enumerate(Config.ENCABEZADOS, 1):
                ancho = anchos.get(encabezado, 15)
                columna_letra = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[columna_letra].width = ancho

            logger.debug("Anchos de columnas ajustados")

        except Exception as e:
            logger.warning(f"Error al ajustar anchos de columnas: {e}")

    @staticmethod
    def generar_nombre_archivo() -> str:
        """
        Genera un nombre de archivo único para el Excel.

        Returns:
            String con el nombre del archivo
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre = f"actas_escaneadas_{timestamp}.xlsx"
        logger.debug(f"Nombre de archivo generado: {nombre}")
        return nombre
